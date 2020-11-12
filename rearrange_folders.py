#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Nov 11 20:20:49 2020

@author: adrian
"""
import os
from shutil import copy2
import openpyxl
import sys

current_folder=sys.argv[1];
os.chdir(current_folder)
folders=['hpc','cortex']
#j=0;
for j in range(2):
    #Create folder brain region
    os.mkdir(folders[j])
    
    #Read excel sheet
    av=[];
    wb_obj = openpyxl.load_workbook(folders[j]+'.xlsx')
    sheet = wb_obj.active
    for row in sheet.iter_rows(max_row=sheet.max_row):
        for cell in row:
            #print(cell.value, end=" ")
            av.append(cell.value)
        #print()
    #Run for all tetrodes
    for i in range(0, len(av), 2):
           #Find tetrode channels
        try:
            os.mkdir(os.getcwd()+'/'+folders[j]+'/'+'Tetrode_'+str(av[i]))
            channels=av[i+1]
            #Split channels, sometimes a comma is converted into a dot.
            try:
                channels=channels.split(',')
            except AttributeError:
                channels=str(channels)
                channels=channels.split('.')
            # Save tetrode .continuous files
            #channels=['14','29','15','16']
            append_str = '100_CH';
            post_str='.continuous'
            channels = [append_str + sub + post_str for sub in channels] 
            
            for x in channels:
                copy2(x,os.getcwd()+'/'+folders[j]+'/'+'Tetrode_'+str(av[i]))
                #os.rename(os.getcwd()+'/'+x,os.getcwd()+'/'+'Tetrode_'+str(av[i])+'/'+x )
            copy2('all_channels.events', os.getcwd()+'/'+folders[j]+'/'+'Tetrode_'+str(av[i]))
            copy2('Continuous_Data.openephys', os.getcwd()+'/'+folders[j]+'/'+'Tetrode_'+str(av[i]))
            copy2('settings.xml', os.getcwd()+'/'+folders[j]+'/'+'Tetrode_'+str(av[i]))
            copy2('messages.events', os.getcwd()+'/'+folders[j]+'/'+'Tetrode_'+str(av[i]))
            copy2('tetrode.prb', os.getcwd()+'/'+folders[j]+'/'+'Tetrode_'+str(av[i]))
        except FileExistsError:
            print('Folder already exists')


